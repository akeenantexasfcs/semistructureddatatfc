#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def process_excel_sheet(df):
    """Transform raw Excel data into structured format"""
    try:
        # Find quality row
        quality_row = None
        for idx, row in df.iterrows():
            if 'Quality' in str(row.iloc[0]):
                quality_row = idx
                break
                
        if quality_row is None:
            raise ValueError("Could not find Quality category")
            
        category = df.iloc[quality_row, 0]
        
        # Initialize structured data list
        structured_data = []
        current_parent = None
        
        # Start processing after header row
        header_row = quality_row + 1
        data_start = header_row + 1
        
        for idx in range(data_start, len(df)):
            row = df.iloc[idx]
            
            # Skip empty rows
            if pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                continue
                
            # Check if this is a parent row (company name)
            if not pd.isna(row.iloc[0]) and pd.isna(row.iloc[1]):
                current_parent = str(row.iloc[0]).strip()
                # Add parent row
                structured_data.append({
                    'Name/Term': current_parent,
                    'LGD': '',
                    '% Used of RR': None,
                    '% Used of AGG': None,
                    'Used': None,
                    'Available': None,
                    'Total Exposure': None,
                    '% TE of RR': None,
                    '% TE of AGG': None,
                    'is_parent': True
                })
            
            # Check if this is a data row
            elif not pd.isna(row.iloc[1]):
                term = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
                structured_data.append({
                    'Name/Term': term,
                    'LGD': str(row.iloc[1]).strip(),
                    '% Used of RR': row.iloc[2],
                    '% Used of AGG': row.iloc[3],
                    'Used': row.iloc[4],
                    'Available': row.iloc[5],
                    'Total Exposure': row.iloc[6],
                    '% TE of RR': row.iloc[7],
                    '% TE of AGG': row.iloc[8],
                    'is_parent': False
                })
                
        return category, pd.DataFrame(structured_data)
    
    except Exception as e:
        st.error(f"Error processing sheet: {str(e)}")
        raise

def create_styled_excel(category, df):
    """Create styled Excel file"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write DataFrame without index
        df_to_write = df.drop('is_parent', axis=1)
        df_to_write.to_excel(writer, index=False, sheet_name='Sheet1')
        
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Add title row
        worksheet.insert_rows(1)
        worksheet['A1'] = category
        worksheet['A1'].font = Font(bold=True)
        
        # Define styles
        yellow_fill = PatternFill(start_color='FFEB9C',
                                end_color='FFEB9C',
                                fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set alignments
        center_aligned = Alignment(horizontal='center', vertical='center')
        right_aligned = Alignment(horizontal='right', vertical='center')
        left_aligned = Alignment(horizontal='left', vertical='center')
        
        # Format cells
        for row_idx, row in enumerate(df.itertuples(), start=3):
            # Format parent rows
            if row.is_parent:
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.fill = yellow_fill
                    cell.font = Font(bold=True)
            else:
                # Name/Term - left aligned
                worksheet.cell(row=row_idx, column=1).alignment = left_aligned
                
                # LGD - center aligned
                worksheet.cell(row=row_idx, column=2).alignment = center_aligned
                
                # Numbers and percentages - right aligned
                for col in range(3, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col)
                    cell.alignment = right_aligned
                    
                    # Format percentages
                    if col in [3, 4, 8, 9]:  # Percentage columns
                        if cell.value:
                            cell.number_format = '0.00%'
                            cell.value = float(cell.value) / 100  # Convert to decimal
                    
                    # Format numbers
                    if col in [5, 6, 7]:  # Number columns
                        if cell.value:
                            cell.number_format = '#,##0'
        
        # Apply borders and set header row
        header_row = worksheet[2]
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.alignment = center_aligned
            
        # Apply borders to all cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 40  # Name/Term
        worksheet.column_dimensions['B'].width = 10  # LGD
        for col in range(3, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    output.seek(0)
    return output

def main():
    st.title("Excel PD Sheet Processor")
    st.write("Upload an Excel workbook with PD sheets to process.")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file:
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names
            
            st.info(f"Found {len(sheet_names)} sheets in the workbook")
            
            selected_sheets = st.multiselect(
                "Select sheets to process",
                sheet_names,
                default=sheet_names[0] if sheet_names else None
            )
            
            if st.button("Process Selected Sheets"):
                for sheet_name in selected_sheets:
                    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
                    category, processed_df = process_excel_sheet(df)
                    
                    # Show preview
                    st.subheader(f"Preview: {sheet_name}")
                    st.dataframe(processed_df.drop('is_parent', axis=1))
                    
                    # Create download button
                    excel_data = create_styled_excel(category, processed_df)
                    st.download_button(
                        label=f"Download {sheet_name} Excel",
                        data=excel_data,
                        file_name=f"{sheet_name}_formatted.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            st.write("Please make sure the Excel file follows the expected format.")

if __name__ == "__main__":
    main()

