#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import pandas as pd
import io
import json
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def safe_numeric_convert(value):
    """Safely convert a value to numeric, returning None if not possible"""
    if pd.isna(value):
        return None
    try:
        if isinstance(value, str):
            value = value.replace(',', '').replace('%', '')
        return float(value)
    except (ValueError, TypeError):
        return None

def process_excel_to_json(df_raw):
    """
    Convert semi-structured Excel to a JSON format that supports:
      {
        "category": "07 - Average Quality",
        "entries": [
          {
            "name": "SLR Property I, LP",
            "term": "...",
            "lgd": "...",
            "metrics": {...}
          },
          {
            "name": "FRIONA INDUSTRIES, L.P.",
            "subCategory": "REVOLVER",
            "entries": [
              {
                "term": "...",
                "lgd": "...",
                "metrics": {...}
              },
              ...
            ]
          }
        ]
      }
    """
    # 1. First row has the category
    category = str(df_raw.iloc[0, 0]).strip()
    
    # 2. Initialize top-level structure
    json_data = {
        "category": category,
        "entries": []
    }
    
    # 3. Variables to keep track of the current parent and subcategory block
    current_parent_entry = None
    current_subcategory_entry = None
    
    # 4. Start iterating from row 2 onward (assuming row 1 is category)
    for i in range(1, len(df_raw)):
        row = df_raw.iloc[i]
        
        # Skip fully empty rows
        if row.isna().all():
            continue

        # Extract columns safely (update indices to match your sheet structure)
        # Example assumption:
        #   Col A: name/term
        #   Col B: LGD
        #   Col C: %RRUsed
        #   Col D: %AGGUsed
        #   Col E: Used
        #   Col F: Available
        #   Col G: Total Exposure
        #   Col H: %TE of RR
        #   Col I: %TE of AGG
        col_name_term = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
        col_lgd = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
        
        # Decide how to interpret each row:
        # -------------------------------------------------------------
        #  (A) Parent row with no metrics (just "SLR Property I, LP")
        #  (B) Simple row with metrics (Term + LGD + rest)
        #  (C) A possible "subCategory" row
        #  (D) Row under a subCategory
        # -------------------------------------------------------------
        
        # This logic is purely an example. You need to define the exact rule(s)
        # that help you differentiate "parent name" vs "subCategory" vs "actual data row."
        
        # (A) Check if row might be a new parent (has a name, but no LGD or other columns)
        # or if col_name_term is obviously a 'headline' row from your data.
        if col_name_term and not col_lgd:
            # Start a new parent
            current_parent_entry = {
                "name": col_name_term,
                # If you also have "subCategory" in the same row, parse it here
                # "subCategory": ...
            }
            json_data["entries"].append(current_parent_entry)
            # Reset subcategory
            current_subcategory_entry = None
            continue
        
        # (B) Check if row might be a subCategory row
        # e.g., if your data has something like "REVOLVER" in col A or a dedicated column.
        # For demonstration, let's assume anything that includes the word "REVOLVER" is a subCategory:
        if "REVOLVER" in col_name_term.upper():
            # Start or continue subCategory block
            current_subcategory_entry = {
                "name": current_parent_entry["name"] if current_parent_entry else "",
                "subCategory": col_name_term,
                "entries": []
            }
            # Append subCategory block to top-level (or to parent's entries if you prefer)
            json_data["entries"].append(current_subcategory_entry)
            continue
        
        # (C) If we have metrics in the row, parse them
        percent_rr_used = safe_numeric_convert(row.iloc[2])
        # If you see there's no numeric data at all, you might skip
        if not any([percent_rr_used, safe_numeric_convert(row.iloc[3]), safe_numeric_convert(row.iloc[4])]):
            # No real data - skip
            continue
        
        # Build the metrics dict
        metrics = {
            "percentRRUsed": percent_rr_used,
            "percentAGGUsed": safe_numeric_convert(row.iloc[3]),
            "used": safe_numeric_convert(row.iloc[4]),
            "available": safe_numeric_convert(row.iloc[5]),
            "totalExposure": safe_numeric_convert(row.iloc[6]),
            "percentTERR": safe_numeric_convert(row.iloc[7]),
            "percentTEAGG": safe_numeric_convert(row.iloc[8])
        }
        
        # (D) Now decide if this row belongs to the parent directly or a subCategory
        entry_dict = {
            "term": col_name_term,
            "lgd": col_lgd,
            "metrics": metrics
        }
        
        if current_subcategory_entry:
            # Add to the subCategory's "entries" array
            current_subcategory_entry["entries"].append(entry_dict)
        elif current_parent_entry:
            # Just store as a "flat" object on the parent
            # If you want each parent to also be "flat," you can do it this way:
            #   parent entries might be: 
            #     { "name": "...", "term": "...", "lgd": "...", "metrics": {...} }
            #   or you can create an "entries" array on the parent
            #   if you expect multiple terms under one parent.
            
            # For example, if your parent can have multiple lines:
            #   current_parent_entry.setdefault("entries", []).append(entry_dict)
            # or if each parent can have only one line, do:
            current_parent_entry.update(entry_dict)
        else:
            # If there's no parent, just push directly at the top level
            json_data["entries"].append({
                "name": "",
                **entry_dict
            })
    
    return json_data


def create_excel_from_json(json_data):
    """
    Create a clean tabular Excel from the JSON data.
    """
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            workbook = writer.book
            worksheet = workbook.create_sheet('Sheet1')
            
            # Delete default "Sheet" created by openpyxl if needed
            default_sheet = workbook["Sheet"]
            workbook.remove(default_sheet)
            
            # Headers
            headers = [
                "Name/Term", "LGD", "% RR Used", "% AGG Used", "Used",
                "Available", "Total Exposure", "% TE of RR", "% TE of AGG"
            ]
            
            # Row 1: Category
            worksheet.cell(row=1, column=1, value=json_data["category"])
            worksheet.cell(row=1, column=1).font = Font(bold=True)
            
            # Row 2: Headers
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            # Styling
            yellow_fill = PatternFill(
                start_color='FFEB9C',
                end_color='FFEB9C',
                fill_type='solid'
            )
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            current_row = 3
            
            def write_data_row(row_num, term, lgd, metrics, indent=0):
                # Term
                worksheet.cell(row=row_num, column=1, value=(" " * indent) + (term or ""))
                # LGD
                worksheet.cell(row=row_num, column=2, value=lgd)
                
                # RR Used
                if metrics.get("percentRRUsed") is not None:
                    cell = worksheet.cell(row=row_num, column=3, value=metrics["percentRRUsed"] / 100)
                    cell.number_format = '0.00%'
                # AGG Used
                if metrics.get("percentAGGUsed") is not None:
                    cell = worksheet.cell(row=row_num, column=4, value=metrics["percentAGGUsed"] / 100)
                    cell.number_format = '0.00%'
                # Used
                if metrics.get("used") is not None:
                    cell = worksheet.cell(row=row_num, column=5, value=metrics["used"])
                    cell.number_format = '#,##0'
                # Available
                if metrics.get("available") is not None:
                    cell = worksheet.cell(row=row_num, column=6, value=metrics["available"])
                    cell.number_format = '#,##0'
                # Total Exposure
                if metrics.get("totalExposure") is not None:
                    cell = worksheet.cell(row=row_num, column=7, value=metrics["totalExposure"])
                    cell.number_format = '#,##0'
                # % TE of RR
                if metrics.get("percentTERR") is not None:
                    cell = worksheet.cell(row=row_num, column=8, value=metrics["percentTERR"] / 100)
                    cell.number_format = '0.00%'
                # % TE of AGG
                if metrics.get("percentTEAGG") is not None:
                    cell = worksheet.cell(row=row_num, column=9, value=metrics["percentTEAGG"] / 100)
                    cell.number_format = '0.00%'
            
            for entry in json_data["entries"]:
                name = entry.get("name", "")
                sub_cat = entry.get("subCategory", "")
                
                # If there's a subCategory, we might show the parent row with fill
                if sub_cat:
                    # Parent with subCategory means the actual data is in entry["entries"]
                    # So first write the row for the parent + subCat, if you want that visible
                    row_cell = worksheet.cell(row=current_row, column=1, value=name + " - " + sub_cat)
                    row_cell.fill = yellow_fill
                    current_row += 1
                    
                    # Now write each sub-entry
                    for sub_entry in entry["entries"]:
                        write_data_row(
                            current_row,
                            sub_entry.get("term"),
                            sub_entry.get("lgd"),
                            sub_entry.get("metrics", {}),
                            indent=2
                        )
                        current_row += 1
                        
                else:
                    # If there's an "entries" array, this parent has multiple lines
                    if "entries" in entry and isinstance(entry["entries"], list):
                        # Write parent name row
                        if name:
                            parent_cell = worksheet.cell(row=current_row, column=1, value=name)
                            parent_cell.fill = yellow_fill
                            current_row += 1
                        
                        # Now sub-entries
                        for sub_entry in entry["entries"]:
                            write_data_row(
                                current_row,
                                sub_entry.get("term"),
                                sub_entry.get("lgd"),
                                sub_entry.get("metrics", {}),
                                indent=2
                            )
                            current_row += 1
                    else:
                        # If it's a "flat" parent with direct metrics
                        # or just a single data row
                        # Write parent name if needed
                        if name and not entry.get("term"):
                            parent_cell = worksheet.cell(row=current_row, column=1, value=name)
                            parent_cell.fill = yellow_fill
                            current_row += 1
                        
                        # If there's also "term" in the same dict, write it
                        if entry.get("term"):
                            write_data_row(
                                current_row,
                                entry["term"],
                                entry.get("lgd", ""),
                                entry.get("metrics", {}),
                                indent=2
                            )
                            current_row += 1
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 40  # Name/Term
            worksheet.column_dimensions['B'].width = 10  # LGD
            for col_idx in range(3, 10):
                worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # Border + alignment
            for row_cells in worksheet.iter_rows(min_row=2, max_row=current_row - 1, min_col=1, max_col=9):
                for cell in row_cells:
                    cell.border = thin_border
                    # Right-align numeric columns
                    if cell.column in [3,4,5,6,7,8,9]:
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left')
            
        output.seek(0)
        return output
    
    except Exception as e:
        raise Exception(f"Error creating Excel file: {str(e)}")

def main():
    st.title("Excel PD Sheet Processor")
    st.write("Upload an Excel workbook with PD sheets to process.")
    
    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx', 'xls'])
    
    if uploaded_file:
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names
            
            st.info(f"Found {len(sheet_names)} sheets in the workbook")
            
            selected_sheets = st.multiselect(
                "Select sheets to process",
                sheet_names,
                default=sheet_names[0] if sheet_names else None
            )
            
            if st.button("Process Selected Sheets"):
                for sheet_name in selected_sheets:
                    try:
                        df_raw = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None)
                        
                        # Convert to desired JSON
                        json_data = process_excel_to_json(df_raw)
                        
                        # Display JSON
                        st.write(f"JSON structure for sheet '{sheet_name}':")
                        st.json(json_data)
                        
                        # Create Excel to download
                        excel_data = create_excel_from_json(json_data)
                        st.download_button(
                            label=f"Download {sheet_name}",
                            data=excel_data,
                            file_name=f"{sheet_name}_formatted.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                    except Exception as e:
                        st.error(f"Error processing sheet '{sheet_name}': {str(e)}")
        except Exception as e:
            st.error(f"Error reading Excel file: {str(e)}")

if __name__ == "__main__":
    main()

