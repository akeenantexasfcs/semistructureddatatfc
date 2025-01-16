#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import streamlit as st
import pandas as pd
import json
import io
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def create_styled_excel(processed_data):
    """Create Excel file with matching styles from the screenshot"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, data in processed_data.items():
            # Convert the data to rows
            rows = []
            for entry in data['entries']:
                # Add the company name row
                rows.append({
                    'Name/Term': entry['name'],
                    'LGD': '',
                    '% RR Used': '',
                    '% AGG Used': '',
                    'Used': '',
                    'Available': '',
                    'Total Exposure': '',
                    '% TE of RR': '',
                    '% TE of AGG': ''
                })
                
                # Add the metrics row
                if entry['metrics']:
                    rows.append({
                        'Name/Term': entry['term'] or '',
                        'LGD': entry['lgd'],
                        '% RR Used': entry['metrics']['percentRRUsed'],
                        '% AGG Used': entry['metrics']['percentAGGUsed'],
                        'Used': entry['metrics']['used'],
                        'Available': entry['metrics']['available'],
                        'Total Exposure': entry['metrics']['totalExposure'],
                        '% TE of RR': entry['metrics']['percentTERR'],
                        '% TE of AGG': entry['metrics']['percentTEAGG']
                    })
                
                # Add sub-entries
                for sub in entry.get('sub_entries', []):
                    rows.append({
                        'Name/Term': sub['term'],
                        'LGD': sub['lgd'],
                        '% RR Used': sub['metrics']['percentRRUsed'],
                        '% AGG Used': sub['metrics']['percentAGGUsed'],
                        'Used': sub['metrics']['used'],
                        'Available': sub['metrics']['available'],
                        'Total Exposure': sub['metrics']['totalExposure'],
                        '% TE of RR': sub['metrics']['percentTERR'],
                        '% TE of AGG': sub['metrics']['percentTEAGG']
                    })
            
            # Create DataFrame and write to Excel
            df = pd.DataFrame(rows)
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Add title row
            worksheet.insert_rows(1)
            worksheet['A1'] = data['category']
            worksheet['A1'].font = Font(bold=True)
            
            # Define styles
            yellow_fill = PatternFill(start_color='FFEB9C',
                                    end_color='FFEB9C',
                                    fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting
            for row_idx, row in enumerate(rows, start=3):  # Start after title and header
                # Highlight company rows
                if row['LGD'] == '':  # This is a company row
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.fill = yellow_fill
                        cell.font = Font(bold=True)
                
                # Apply number formatting
                if row['LGD']:  # This is a data row
                    # Format percentages
                    for col in [3, 4, 8, 9]:  # Percentage columns
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.number_format = '0.00%'
                    
                    # Format numbers
                    for col in [5, 6, 7]:  # Number columns
                        cell = worksheet.cell(row=row_idx, column=col)
                        cell.number_format = '#,##0'
            
            # Apply borders to all cells
            for row in worksheet.iter_rows(min_row=1, max_row=len(rows)+2):
                for cell in row:
                    cell.border = thin_border
            
            # Adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                worksheet.column_dimensions[column].width = min(max_length + 2, 30)

    output.seek(0)
    return output

def main():
    st.title("Excel PD Sheet Processor")
    # ... rest of the main function remains the same ...
    # Just replace the Excel export part with:
    
    excel_data = create_styled_excel(processed_data)
    st.download_button(
        label="Download Formatted Excel",
        data=excel_data,
        file_name="formatted_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

